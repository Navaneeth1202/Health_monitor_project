import os

from datetime import datetime

from openpyxl import Workbook, load_workbook

from config import excel_file

# =========================================================
# CREATE EXCEL FILE
# =========================================================

if not os.path.exists(excel_file):

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Health Logs"

    sheet.append([
        "Timestamp",
        "Systolic BP",
        "Diastolic BP",
        "Heart Rate",
        "Activity",
        "Status",
        "Reason"
    ])

    workbook.save(excel_file)

# =========================================================
# SAVE DATA
# =========================================================

def save_health_data(
    bp_sys,
    bp_dia,
    heart_rate,
    activity,
    status,
    reason
):

    workbook = load_workbook(excel_file)

    sheet = workbook.active

    timestamp = datetime.now().strftime(
        "%Y-%m-%d %H:%M:%S"
    )

    sheet.append([
        timestamp,
        bp_sys,
        bp_dia,
        heart_rate,
        activity,
        status,
        reason
    ])

    workbook.save(excel_file)