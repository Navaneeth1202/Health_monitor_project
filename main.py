<<<<<<< HEAD
import time

from bp_generator import generate_bp

from heart_rate_generator import (
    generate_heart_rate
)

from motion_detector import detect_motion

from excel_database import save_health_data

from bp_alerts import (
    low_bp_alert,
    high_bp_alert
)

from workout import workout_confirmation

from emergency import emergency_sos

# =========================================================
# GLOBAL VARIABLE
# =========================================================

workout_confirmed = False

# =========================================================
=======
import random
import time
import threading
import winsound
import os

from datetime import datetime
from twilio.rest import Client
from openpyxl import Workbook, load_workbook

# =========================================================
# TWILIO CONFIGURATION
# =========================================================

account_sid = "YOUR_ACCOUNT_SID"
auth_token = "YOUR_AUTH_TOKEN"

twilio_number = "+1XXXXXXXXXX"

client = Client(account_sid, auth_token)

# =========================================================
# CONTACT DETAILS
# =========================================================

primary_contact = "+919952962137"

ambulance_number = "108"

# =========================================================
# GLOBAL VARIABLES
# =========================================================

sos_acknowledged = False

workout_confirmed = False

# =========================================================
# EXCEL DATABASE SETUP
# =========================================================

excel_file = "health_monitor_data.xlsx"

if not os.path.exists(excel_file):

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Health Logs"

    sheet.append([
        "Timestamp",
        "Systolic BP",
        "Diastolic BP",
        "Heart Rate",
        "Activity",
        "Status",
        "Reason"
    ])

    workbook.save(excel_file)

# =========================================================
# REALISTIC BP GENERATOR
# =========================================================

def generate_bp():

    systolic = random.randint(90, 170)

    diastolic = random.randint(60, 100)

    # realistic BP gap
    if diastolic >= systolic:

        diastolic = systolic - random.randint(30, 50)

    return systolic, diastolic

# =========================================================
# RANDOM HEART RATE GENERATOR
# =========================================================

def generate_heart_rate():

    return random.randint(55, 170)

# =========================================================
# SMART MOTION DETECTION
# =========================================================

def detect_motion(bp_sys, bp_dia, heart_rate):

    # FALL CONDITION
    if (

        (
            (bp_sys < 90 or bp_dia < 60)
            or
            (bp_sys > 140 or bp_dia > 90)
        )

        and heart_rate > 130
    ):

        return "FALL"

    # WORKOUT CONDITION
    elif heart_rate > 120:

        return random.choice([
            "RUNNING",
            "WORKOUT"
        ])

    # NORMAL CONDITION
    else:

        return random.choice([
            "RESTING",
            "WALKING"
        ])

# =========================================================
# SAVE DATA TO EXCEL
# =========================================================

def save_health_data(

    bp_sys,
    bp_dia,
    heart_rate,
    activity,
    status,
    reason
):

    workbook = load_workbook(excel_file)

    sheet = workbook.active

    timestamp = datetime.now().strftime(
        "%Y-%m-%d %H:%M:%S"
    )

    sheet.append([

        timestamp,

        bp_sys,

        bp_dia,

        heart_rate,

        activity,

        status,

        reason
    ])

    workbook.save(excel_file)

# =========================================================
# REAL SMS FUNCTION
# =========================================================

def send_sms(number):

    try:

        client.messages.create(

            body="🚨 MEDICAL EMERGENCY 🚨",

            from_=twilio_number,

            to=number
        )

        print("\n📩 MEDICAL EMERGENCY SMS SENT")

    except Exception as e:

        print("\n❌ SMS FAILED")

        print(e)

# =========================================================
# CONTINUOUS EMERGENCY SOUND
# =========================================================

def beep_alarm():

    global sos_acknowledged

    while not sos_acknowledged:

        winsound.Beep(1200, 500)

        winsound.Beep(800, 500)

# =========================================================
# AMBULANCE CALL
# =========================================================

def call_ambulance():

    print("\n🚑 CALLING AMBULANCE...")

# =========================================================
# LOW BP ALERT
# =========================================================

def low_bp_alert():

    print("\n⚠ LOW BP DETECTED ⚠")

    for i in range(5):

        winsound.Beep(1000, 800)

        time.sleep(0.3)

    print("\n⬅ ➡ ⬆ ⬇")

    print("Press D to confirm safe condition")

    response = input("Button: ")

    if response.upper() == "D":

        print("\n✅ SAFE CONDITION CONFIRMED")

# =========================================================
# HIGH BP ALERT
# =========================================================

def high_bp_alert():

    print("\n⚠ HIGH BP DETECTED ⚠")

    for i in range(5):

        winsound.Beep(1200, 800)

        time.sleep(0.3)

    print("\n⬅ ➡ ⬆ ⬇")

    print("Press D to confirm safe condition")

    response = input("Button: ")

    if response.upper() == "D":

        print("\n✅ SAFE CONDITION CONFIRMED")

# =========================================================
# WORKOUT CONFIRMATION
# =========================================================

def workout_confirmation():

    print("\n🏃 HIGH HEART RATE DURING ACTIVITY")

    winsound.Beep(900, 500)

    print("\n⬅ ➡ ⬆ ⬇")

    print("Press D to confirm workout")

    response = input("Button: ")

    if response.upper() == "D":

        print("\n✅ WORKOUT CONFIRMED")

# =========================================================
# EMERGENCY SOS
# =========================================================

def emergency_sos():

    global sos_acknowledged

    sos_acknowledged = False

    # Start emergency sound
    alarm_thread = threading.Thread(
        target=beep_alarm
    )

    alarm_thread.start()

    # Send emergency SMS
    send_sms(primary_contact)

    # Wait for user response
    start_time = time.time()

    while time.time() - start_time < 120:

        print("\n⬅ ➡ ⬆ ⬇")

        print("Press D to stop emergency alert")

        response = input("Button: ")

        if response.upper() == "D":

            sos_acknowledged = True

            print("\n✅ EMERGENCY ALERT STOPPED")

            return

    # No response
    sos_acknowledged = True

    call_ambulance()

# =========================================================
>>>>>>> fda21c194f0d77b636d610189facf4ca264c31c2
# MAIN LOOP
# =========================================================

while True:

    print("\n================================================")

    print("            AI HEALTH MONITOR")

    print("================================================")

<<<<<<< HEAD
=======
    # Generate values
>>>>>>> fda21c194f0d77b636d610189facf4ca264c31c2
    bp_sys, bp_dia = generate_bp()

    heart_rate = generate_heart_rate()

<<<<<<< HEAD
    activity = detect_motion(
        bp_sys,
        bp_dia,
        heart_rate
    )

=======
    # Smart activity detection
    activity = detect_motion(

        bp_sys,

        bp_dia,

        heart_rate
    )

    # Display
>>>>>>> fda21c194f0d77b636d610189facf4ca264c31c2
    print(f"\n🩸 BP: {bp_sys}/{bp_dia}")

    print(f"❤ HEART RATE: {heart_rate} BPM")

    print(f"🏃 ACTIVITY: {activity}")

<<<<<<< HEAD
=======
    # =====================================================
    # EMERGENCY CONDITION
    # =====================================================

>>>>>>> fda21c194f0d77b636d610189facf4ca264c31c2
    emergency_condition = (

        (
            (bp_sys < 90 or bp_dia < 60)
            or
            (bp_sys > 140 or bp_dia > 90)
        )

        and heart_rate > 130

        and activity == "FALL"
    )

<<<<<<< HEAD
=======
    # =====================================================
    # STATUS DETECTION
    # =====================================================

>>>>>>> fda21c194f0d77b636d610189facf4ca264c31c2
    status = "NORMAL"

    reason = "Vitals Stable"

    if emergency_condition:

        status = "EMERGENCY"

        reason = "Fall + Abnormal BP/Heart Rate"

    elif bp_sys > 140 or bp_dia > 90:

        status = "HIGH BP"

        reason = "Blood Pressure High"

    elif bp_sys < 90 or bp_dia < 60:

        status = "LOW BP"

        reason = "Blood Pressure Low"

    elif heart_rate > 110:

        status = "WARNING"

        reason = "Heart Rate High"

    print(f"\n📋 STATUS: {status}")

    print(f"📌 REASON: {reason}")

<<<<<<< HEAD
    save_health_data(
        bp_sys,
        bp_dia,
        heart_rate,
        activity,
        status,
=======
    # =====================================================
    # SAVE TO EXCEL
    # =====================================================

    save_health_data(

        bp_sys,

        bp_dia,

        heart_rate,

        activity,

        status,

>>>>>>> fda21c194f0d77b636d610189facf4ca264c31c2
        reason
    )

    print("\n💾 DATA SAVED TO EXCEL")

    # =====================================================
<<<<<<< HEAD
    # WORKOUT MODE
=======
    # SMART WORKOUT DETECTION
>>>>>>> fda21c194f0d77b636d610189facf4ca264c31c2
    # =====================================================

    if activity in ["RUNNING", "WORKOUT"]:

<<<<<<< HEAD
=======
        # Ask only once
>>>>>>> fda21c194f0d77b636d610189facf4ca264c31c2
        if not workout_confirmed:

            workout_confirmation()

            workout_confirmed = True

<<<<<<< HEAD
        if (
=======
        # Dangerous condition during workout
        if (

>>>>>>> fda21c194f0d77b636d610189facf4ca264c31c2
            (
                (bp_sys < 90 or bp_dia < 60)
                or
                (bp_sys > 160 or bp_dia > 100)
            )
<<<<<<< HEAD
            or
            heart_rate > 165
        ):

            print(
                "\n🚨 ABNORMAL CONDITION DURING WORKOUT 🚨"
            )

            emergency_sos()

=======

            or

            heart_rate > 165
        ):

            print("\n🚨 ABNORMAL CONDITION DURING WORKOUT 🚨")

            emergency_sos()

    # Reset workout mode
>>>>>>> fda21c194f0d77b636d610189facf4ca264c31c2
    else:

        workout_confirmed = False

    # =====================================================
    # LOW BP ALERT
    # =====================================================

    if (
<<<<<<< HEAD
        (bp_sys < 90 or bp_dia < 60)
=======

        (bp_sys < 90 or bp_dia < 60)

        and heart_rate < 130

>>>>>>> fda21c194f0d77b636d610189facf4ca264c31c2
        and activity == "RESTING"
    ):

        low_bp_alert()

    # =====================================================
    # HIGH BP ALERT
    # =====================================================

    elif (
<<<<<<< HEAD
        (bp_sys > 140 or bp_dia > 90)
=======

        (bp_sys > 140 or bp_dia > 90)

        and heart_rate < 130

>>>>>>> fda21c194f0d77b636d610189facf4ca264c31c2
        and activity == "RESTING"
    ):

        high_bp_alert()

    # =====================================================
<<<<<<< HEAD
    # FALL EMERGENCY
=======
    # EMERGENCY DETECTION
>>>>>>> fda21c194f0d77b636d610189facf4ca264c31c2
    # =====================================================

    elif emergency_condition:

        emergency_sos()

<<<<<<< HEAD
    print("\n⏳ NEXT CHECK IN 10 SECONDS...\n")

    time.sleep(10)
=======
    # =====================================================
    # NEXT CHECK
    # =====================================================

    print("\n⏳ NEXT CHECK IN 10 SECONDS...\n")

    # TEST MODE
    time.sleep(10)

    # REAL MODE
    # time.sleep(300)
>>>>>>> fda21c194f0d77b636d610189facf4ca264c31c2
